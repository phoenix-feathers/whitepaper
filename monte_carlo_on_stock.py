# Cell 1 for Monte Carlo on Stocks
# OPTIONAL CELL
from google.colab import drive
drive.mount('/content/drive')


# Cell 2 for Monte Carlo on Stocks
# function to change the returns into new set of returns
def return_returns(rows, returns, add):
  temp_returns = []
  if (len(returns)==(rows-1)):
    temp_returns.append(returns[0])
  for i in range(1, len(returns)):
    temp_returns.append(returns[i])
  temp_returns.append(add)
  return temp_returns


# function to calculate mean and standard deviation
def mean_and_sigma (rows, returns):
  limit = len(returns)
  mu = 0
  for i in range(limit):
    mu += returns[i]
  mu/=rows
  sigma = 0
  for i in range(limit):
    sigma+=((returns[i]-mu)**2)
  sigma/=rows
  sigma = sigma**0.5
  return mu,sigma

# function for next stock price and final return
def values(returns, rows, so,timestep):
  mu, sigma = mean_and_sigma(rows,returns)
  e = random.uniform(-2,2)
  return_ = mu/timestep + sigma*e*(timestep**-0.5)
  returns_ = return_returns(rows,returns,return_)
  s = so*(return_ + 1)
  return returns_,s
  
  
  # Cell 3 for Monte Carlo on Stocks
# plots
def plotting(opens,closes,paths,rows,iters, dates, prev_dates):
  
  # plotting historical data
  f = plt.figure()
  f.set_figwidth(20)
  f.set_figheight(10)
  plt.plot(prev_dates, closes, label = 'Closes')
  plt.plot(prev_dates, opens, label = 'Opens')
  plt.title('Historical data', fontweight = 'bold')
  plt.xlabel('Timesteps')
  plt.ylabel('Stock closing price')
  plt.legend()
  plt.show()

  # plotting the paths generated and making a list of the price at the end of the time period
  # also making the list for an average plot
  finals = []
  avg_path = []
  f = plt.figure()
  f.set_figwidth(20)
  f.set_figheight(10)
  for i in range(rows+1):
    avg_path.append(0)
  for y in paths:
    plt.plot(dates,y)
    finals.append(y[rows])
    for i in range(rows+1):
      avg_path[i] += y[i]
  for i in range(rows+1):
    avg_path[i] /= iters
  plt.title('Generated paths', fontweight = 'bold')
  plt.xlabel('Timesteps')
  plt.ylabel('Stock closing price')
  plt.show()

  #plotting the cumulative average of the final stock prices
  cumm = 0
  xx = []
  yy = []
  for i in range(len(finals)):
    cumm = (cumm*i + finals[i])/(i+1)
    xx.append(i)
    yy.append(cumm)
  f = plt.figure()
  f.set_figwidth(10)
  f.set_figheight(5)
  plt.plot(xx,yy)
  plt.title('Cumulative average of final stock price', fontweight = 'bold')
  plt.xlabel('Trails')
  plt.ylabel('Cumulative average of final stock price')
  plt.show()

  # plotting the average path
  f = plt.figure()
  f.set_figwidth(20)
  f.set_figheight(10)
  plt.plot(dates,avg_path)
  plt.title('Estimated path', fontweight = 'bold')
  plt.xlabel('Timesteps')
  plt.ylabel('Stock closing price')
  plt.show()

  # finding the maximum and minimum stock price
  maximum = max(finals)
  minimum = min(finals)

  # histogram plot for final stock prices
  a = numpy.arange(round(minimum) - 1, round(maximum))
  fig,ax = plt.subplots()
  ax.hist(finals,bins = a)
  plt.title('Histogram of the final stock prices', fontweight = 'bold')
  plt.show()
  
  # returns
  return avg_path[rows], minimum, maximum, avg_path


# Cell 4 for Monte Carlo on Stocks
# number of trading days in a month is being taken as 22
month = 22

# monte carlo body
def montecarlo (sheetnum, iters, timestep):

  # opening the sheet needed to open
  sheet = wb[sheets[sheetnum]]
  # the number of rows is the number of entries for a month + the heading
  # the columns have contents in the order Dates, Series, Open, High, Low, Prev. Close, LTP, Close, VWAP, 52W H, 52W L, Volume, Value, No. of Trades
  rows = month*timestep + 1

  # getting the next trading dates to happen for a month
  dates = []
  counter = 0
  date_temp = sheet.cell(rows,1).value
  daychange = datetime.timedelta(days=1/timestep)
  dates.append(date_temp)
  while True:
    date_temp = date_temp + daychange
    if date_temp.isoweekday() < 6:
      dates.append(date_temp)
      counter+=1
    if counter == month*timestep:
      break

  # getting open prices, close prices and dates
  opens = []
  closes = []
  prev_dates = []
  for i in range(2, rows + 1):
    opens.append(sheet.cell(i,3).value)
    closes.append(sheet.cell(i,8).value)
    prev_dates.append(sheet.cell(i,1).value)

  # monte carlo simulation
  paths = []
  returns_ = []
  for i in range(rows-2):
    returns_.append((closes[i+1]/closes[i])-1)
  for i in range(iters):
    path = []
    random.seed()
    returns = returns_
    path.append(closes[rows-2])
    for j in range(rows-1):
      returns, s = values(returns,rows-1, path[j], timestep)
      path.append(s)
    paths.append(path)

  # calling the plot function
  estimate, min, max, avg_path = plotting(opens,closes,paths, rows-1, iters, dates, prev_dates)

  # final print statements
  print(f'The estimated price of stock at the end of month is', end = " ")
  print("%.2f"%estimate + " INR.")
  print(f'The maximum estimated price of the stock at the end of month is', end = " ")
  print("%.2f"%max + " INR.")
  print(f'The minimum estimated price of the stock at the end of month is', end = " ")
  print("%.2f"%min + " INR.")


  # return statement
  paths.append(avg_path)
  paths.append(dates)
  return paths


# Cell 5 for Monte Carlo on Stocks
# required imports
import openpyxl as op
import datetime
import random
import matplotlib.pyplot as plt
import numpy

# opening workbook
wb = op.load_workbook(input("Enter path of the file: "))
sheets = wb.sheetnames
wb_report = op.Workbook()

# conditions
iters = int(input("Enter number of trails to be done: "))
timestep = int(input("Enter number of times the note of stock price is taken in a day: "))

# Monte Carlo simulation
paths = montecarlo(0, iters, timestep)

# saving in a new spreadsheet
name = " Report"
sheet_name = sheets[0] + name
wb_report['Sheet'].title=sheet_name
sheet = wb_report.active
headings = []
c = "Trail "
for i in range(iters,0,-1):
  x = c + str(i)
  headings.append(x)
headings.append("Average path")
headings.append("Dates")
entries = 22*timestep + 1
counter = 0
for i in range(iters+1, -1, -1):
  counter += 1
  sheet.cell(row=1, column = counter).value=headings[i]
  for j in range(entries):
    sheet.cell(row = j+2, column = counter).value=paths[i][j]
name = input("Enter the path to store the data: ")
wb_report.save(name)
